"""Export `analysis/kpis.json` to a shareable Excel workbook.

    python -m scripts.export_kpis_excel

Campaign 1 only. The workbook is for a reader who will not open the JSON, so
every caveat carried in `kpis.json` is written into a visible cell rather than
being lost in the numeric conversion: a rate never appears without its
denominator, and a flagged figure never appears without its flag on the same
sheet, adjacent to the rows it qualifies.

Sheet order follows the report: Overview, one sheet per research question,
Reliability, Efficiency, then what was deliberately not computed.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
CAVEAT_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(bold=True)
CAVEAT_FONT = Font(bold=True, color="9C5700")
TITLE_FONT = Font(bold=True, size=13)
PERCENT_FORMAT = "0.0%"
MAX_WIDTH = 90


def title(sheet: Worksheet, text: str) -> None:
    """Write a sheet title on row 1."""
    sheet.cell(row=1, column=1, value=text).font = TITLE_FONT


def header(sheet: Worksheet, row: int, labels: list[str]) -> int:
    """Write a bold, filled header row. Returns the next free row."""
    for column, label in enumerate(labels, start=1):
        cell = sheet.cell(row=row, column=column, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    return row + 1


def caveat(sheet: Worksheet, row: int, text: str) -> int:
    """Write a highlighted caveat so a flag cannot be mistaken for decoration."""
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = CAVEAT_FONT
    cell.fill = CAVEAT_FILL
    return row + 1


def write_rate(sheet: Worksheet, row: int, column: int, r: dict[str, Any]) -> None:
    """Write a rate as three cells: n, denominator, formatted percentage.

    A rate with no denominator writes its explanatory note instead of a
    number, so an absent measure reads as absent rather than as zero.
    """
    if r.get("rate") is None:
        sheet.cell(row=row, column=column, value=r.get("note") or "N/A")
        return
    sheet.cell(row=row, column=column, value=r["n"])
    sheet.cell(row=row, column=column + 1, value=r["of"])
    cell = sheet.cell(row=row, column=column + 2, value=r["rate"])
    cell.number_format = PERCENT_FORMAT


def finish(sheet: Worksheet, freeze: str = "A4") -> None:
    """Freeze the header and size every column to its content."""
    sheet.freeze_panes = freeze
    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                widths[cell.column] = max(
                    widths.get(cell.column, 0), len(str(cell.value))
                )
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = min(
            width + 2, MAX_WIDTH
        )


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------


def sheet_overview(book: Workbook, k: dict[str, Any]) -> None:
    """Plain-text framing: what this campaign is, and how to read the numbers."""
    sheet = book.create_sheet("Overview")
    meta = k["meta"]
    title(sheet, "LUXCAL - Campaign 1 (batch 1 of 2) - KPI results")
    row = 3
    row = header(sheet, row, ["Item", "Value"])

    for label, value in [
        ("Campaign", "Campaign 1 (batch 1) of 2"),
        ("Runs", meta["n_runs"]),
        (
            "Design",
            f"{meta['n_cases']} cases x {len(meta['variants'])} variants "
            f"x {meta['n_replicates']} replicates = {meta['n_runs']} runs",
        ),
        ("Cases", "case_004 - case_014 (development fixtures 001-003 excluded)"),
        ("Variants", ", ".join(meta["variants"])),
        ("Approximate API cost", "~$8 USD across all 165 runs"),
        (
            "Research questions",
            "4 - RQ1 architecture vs prompt, RQ2 calibration bias, "
            "RQ3 rules vs judgement, RQ4 loop behaviour",
        ),
        ("Generation model", "claude-sonnet-4-6"),
        ("Judge model", "claude-opus-4-6"),
        ("Source", "analysis/kpis.json, produced by scripts/kpi.py"),
    ]:
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    row = caveat(sheet, row, "HOW TO READ THIS WORKBOOK")
    for line in [
        "Every rate is given as three columns - n, denominator, percentage. "
        "No percentage appears without the counts behind it.",
        "Caveats are highlighted rows placed beside the figures they qualify. "
        "A highlighted row is not decoration; it changes how the number may be read.",
        "Three figures must never be quoted without their flag: RQ4 loop effect "
        "(effective n = 1), Reliability evidence rate (contaminated), and "
        "Efficiency cost-per-improvement (illustrative only).",
    ]:
        sheet.cell(row=row, column=1, value=line)
        row += 1

    row += 1
    row = caveat(sheet, row, "CAMPAIGN 2 (planned)")
    for line in [
        "Adds the minus_calibration and minus_profiler ablation variants.",
        "Corrects the Agent 1 evidence-attribution defect that contaminates the "
        "evidence rate reported here.",
        "Will be appended in this same workbook format for direct comparison.",
        "See the 'Deferred - Not Computed' sheet for what Campaign 1 cannot answer.",
    ]:
        sheet.cell(row=row, column=1, value=line)
        row += 1
    finish(sheet)


def sheet_rq1(book: Workbook, k: dict[str, Any]) -> None:
    """The full vs baseline contrast, at the precision stored in the JSON."""
    sheet = book.create_sheet("RQ1 - Architecture vs Prompt")
    r1 = k["rq1"]
    title(sheet, "RQ1 - Does the architecture beat a single prompt?  (full vs baseline)")
    row = 3

    measures = [
        ("2.1 Locus-position inconsistency", "locus_position_inconsistency_2_1"),
        ("2.2 Band-position incoherence", "band_position_incoherence_2_2"),
        ("2.3 Locus match", "locus_match_2_3"),
        ("1.1 Ceiling breach", "ceiling_breach_1_1"),
        ("1.2 Misdeclaration", "misdeclaration_1_2"),
    ]
    labels = ["Variant", "Runs", "Concepts"]
    for name, _ in measures:
        labels += [f"{name} n", f"{name} of", f"{name} %"]
    row = header(sheet, row, labels)

    for variant in k["meta"]["variants"]:
        entry = r1["per_variant"][variant]
        sheet.cell(row=row, column=1, value=variant)
        sheet.cell(row=row, column=2, value=entry["n_runs"])
        sheet.cell(row=row, column=3, value=entry["n_concepts"])
        column = 4
        for _, key in measures:
            write_rate(sheet, row, column, entry[key])
            column += 3
        row += 1

    row += 1
    row = caveat(sheet, row, "CAVEATS - these change how the columns above may be read")
    for text in [
        f"2.1: {r1['notes']['2_1']}",
        "2.2: claimed_visibility is compared against the DECLARED ai_position's "
        "native visibility (BACKSTAGE=LOW, ADVISOR_MEDIATED=MEDIUM, "
        "CLIENT_FACING=HIGH), not against the locus. 2.1 and 2.2 are separate "
        "coherence checks and do not overlap.",
        f"2.3 and 1.1 for baseline: {r1['notes']['2_3_and_1_1_baseline']}",
        "1.2: 'N/A' for minus_critic and baseline - neither runs a Critic, so no "
        "misdeclaration flag exists. This is a structural absence, not a zero.",
    ]:
        sheet.cell(row=row, column=1, value=text)
        row += 1
    finish(sheet)


def sheet_rq2(book: Workbook, k: dict[str, Any]) -> None:
    """Descriptive calibration distributions."""
    sheet = book.create_sheet("RQ2 - Invisibility Calibration")
    r2 = k["rq2"]
    title(sheet, "RQ2 - Is calibration biased toward invisibility?  (descriptive)")
    row = 3

    bands = r2["band_distribution_5_1"]
    row = header(sheet, row, ["5.1 Band distribution", "Band", "Runs", "Denominator"])
    for axis in ("visibility", "intensity"):
        for band, n in bands[axis].items():
            for column, value in enumerate(
                [axis, band, n, bands["denominator"]], start=1
            ):
                sheet.cell(row=row, column=column, value=value)
            row += 1

    row += 1
    surv = r2["locus_survival_5_2"]
    row = header(sheet, row, ["5.2 Locus survival", "Value"])
    for label, value, as_percent in [
        ("Mean loci surviving", surv["mean_loci_surviving"], False),
        ("Grid size", surv["grid_size"], False),
        ("Mean fraction surviving", surv["mean_fraction_surviving"], True),
        ("Denominator (runs that proceeded)", surv["denominator_runs_proceeding"], False),
    ]:
        sheet.cell(row=row, column=1, value=label)
        cell = sheet.cell(row=row, column=2, value=value)
        if as_percent:
            cell.number_format = PERCENT_FORMAT
        row += 1

    row += 1
    excl = r2["exclusion_reason_5_3"]
    row = header(sheet, row, ["5.3 Exclusion reason", "Count", "Denominator", "Share"])
    for reason, n in excl["counts"].items():
        for column, value in enumerate(
            [reason, n, excl["denominator_exclusions"]], start=1
        ):
            sheet.cell(row=row, column=column, value=value)
        cell = sheet.cell(row=row, column=4, value=n / excl["denominator_exclusions"])
        cell.number_format = PERCENT_FORMAT
        row += 1

    row += 1
    freq = r2["locus_frequency_5_4"]
    row = header(sheet, row, ["5.4 Selected locus", "Runs", "Denominator", "Share"])
    for locus, n in freq["counts"].items():
        for column, value in enumerate([locus, n, freq["denominator"]], start=1):
            sheet.cell(row=row, column=column, value=value)
        cell = sheet.cell(row=row, column=4, value=n / freq["denominator"])
        cell.number_format = PERCENT_FORMAT
        row += 1

    row += 1
    row = caveat(sheet, row, "CAVEATS")
    for text in [
        f"5.4: {freq['note']}",
        "5.1, 5.2, 5.3: baseline runs are excluded - that variant produces no "
        "calibration, so it has no bands, no surviving loci and no exclusions.",
    ]:
        sheet.cell(row=row, column=1, value=text)
        row += 1
    finish(sheet)


def sheet_rq3(book: Workbook, k: dict[str, Any]) -> None:
    """The 9/2 agreement split, with each disagreement named and directed."""
    sheet = book.create_sheet("RQ3 - Rules vs Judgement")
    r3 = k["rq3"]
    title(sheet, "RQ3 - Rules vs model judgement for calibration?  (full vs llm_bands)")
    row = 3

    row = header(sheet, row, ["Measure", "n", "of", "%"])
    for label, value in [
        ("Per-case band agreement", r3["band_agreement"]),
        ("1.1 Ceiling breach - full", r3["ceiling_breach_1_1_full"]),
        ("1.1 Ceiling breach - llm_bands", r3["ceiling_breach_1_1_llm_bands"]),
    ]:
        sheet.cell(row=row, column=1, value=label)
        write_rate(sheet, row, 2, value)
        row += 1

    row += 1
    row = header(
        sheet, row, [f"Agreeing cases ({len(r3['agreeing_cases'])})", "Outcome"]
    )
    for case_id in r3["agreeing_cases"]:
        sheet.cell(row=row, column=1, value=case_id)
        sheet.cell(row=row, column=2, value="identical bands")
        row += 1

    row += 1
    row = header(
        sheet,
        row,
        [
            f"Disagreeing cases ({len(r3['disagreeing_cases'])})",
            "full (deterministic)",
            "llm_bands",
            "Direction of disagreement",
        ],
    )
    for entry in r3["disagreeing_cases"]:
        for column, value in enumerate(
            [
                entry["case_id"],
                entry["full_deterministic"],
                entry["llm_bands"],
                entry["direction"],
            ],
            start=1,
        ):
            sheet.cell(row=row, column=column, value=value)
        row += 1

    row += 1
    row = caveat(sheet, row, "CAVEATS")
    for text in [
        r3["note"],
        "Both disagreements run the same way: the model set a STRICTER ceiling "
        "than the arithmetic, never a looser one. With n = 2 this is an "
        "observation, not a demonstrated pattern.",
    ]:
        sheet.cell(row=row, column=1, value=text)
        row += 1
    finish(sheet)


def sheet_rq4(book: Workbook, k: dict[str, Any]) -> None:
    """Loop behaviour, with the n=1 flag placed before any number is read."""
    sheet = book.create_sheet("RQ4 - Loop Behaviour")
    r4 = k["rq4"]
    title(sheet, "RQ4 - Does the loop help or over-smooth?   QUALITATIVE ONLY")
    row = 3

    row = caveat(
        sheet,
        row,
        "EFFECTIVE n = 1 FOR ANY LOOP-EFFECT CLAIM - QUALITATIVE EVIDENCE ONLY. "
        "One revision fired across 33 full runs. No rate may be inferred from "
        "the tables below.",
    )
    row += 1

    row = header(
        sheet,
        row,
        [
            "Variant",
            "3.1 First-pass accept n",
            "of",
            "%",
            "3.3 Escalation n",
            "of",
            "%",
            "3.2 Iteration distribution",
        ],
    )
    for variant in k["meta"]["variants"]:
        entry = r4["per_variant"][variant]
        sheet.cell(row=row, column=1, value=variant)
        if "note" in entry:
            sheet.cell(row=row, column=2, value=entry["note"])
        else:
            write_rate(sheet, row, 2, entry["first_pass_acceptance_3_1"])
            write_rate(sheet, row, 5, entry["escalation_3_3"])
        sheet.cell(
            row=row, column=8, value=json.dumps(entry["iteration_distribution_3_2"])
        )
        row += 1

    row += 1
    quality = r4["quality_improvement_3_5"]
    row = caveat(
        sheet,
        row,
        f"3.5 QUALITY IMPROVEMENT - EFFECTIVE n = {quality['effective_n']}. "
        f"{quality['note']}",
    )
    row = header(
        sheet,
        row,
        [
            "Case",
            "Replicate",
            "Before verdict",
            "Before misdeclared",
            "Before violations",
            "Directive",
            "After verdict",
            "After misdeclared",
            "After violations",
        ],
    )
    for case in quality["cases"]:
        before, after = case["before"], case["after"]
        for column, value in enumerate(
            [
                case["case_id"],
                case["replicate"],
                before["verdict"],
                before["misdeclared"],
                ", ".join(before["violations"]) or "none",
                before["directive"],
                after["verdict"],
                after["misdeclared"],
                ", ".join(after["violations"]) or "none",
            ],
            start=1,
        ):
            sheet.cell(row=row, column=column, value=value)
        row += 1

    row += 1
    row = caveat(sheet, row, f"3.4 OSCILLATION - NOT COMPUTABLE. {r4['oscillation_3_4']['note']}")
    sheet.cell(
        row=row,
        column=1,
        value="Over-smoothing is therefore neither observed nor ruled out: the "
        "loop fired once in 33 full runs, so this campaign cannot distinguish "
        "'the loop does not over-smooth' from 'the loop was rarely exercised'.",
    )
    finish(sheet, freeze="A7")


def sheet_reliability(book: Workbook, k: dict[str, Any]) -> None:
    """Gate behaviour and Agent 1 stability, with the contamination flag."""
    sheet = book.create_sheet("Reliability")
    rel = k["reliability"]
    dims = ["D1", "D2", "D3", "D4", "D5"]
    title(sheet, "Reliability - gate behaviour and Agent 1 stability")
    row = 3

    gate = rel["gate_refusal_1_4"]
    row = header(sheet, row, ["1.4 Gate refusal", "n", "of", "%", "Cases refused"])
    sheet.cell(row=row, column=1, value="Refusal rate")
    write_rate(sheet, row, 2, gate)
    sheet.cell(row=row, column=5, value=json.dumps(gate["cases_refused"]))
    row += 1
    row = caveat(sheet, row, f"1.4: {gate['note']}")
    row = caveat(
        sheet,
        row,
        "1.4: every refusal is case_010 (Hermes) - an authentication and "
        "counterfeiting problem, correctly identified as not a personalisation "
        "problem. No other case was refused in any run.",
    )

    row += 1
    row = header(
        sheet,
        row,
        [
            "4.1 Agent 1 position agreement",
            "Cases unanimous n",
            "of",
            "%",
            "Mean modal share",
            "Runs per case",
        ],
    )
    for dim in dims:
        entry = rel["agent1_position_agreement_4_1"][dim]
        sheet.cell(row=row, column=1, value=dim)
        write_rate(sheet, row, 2, entry["cases_unanimous"])
        cell = sheet.cell(row=row, column=5, value=entry["mean_modal_share"])
        cell.number_format = PERCENT_FORMAT
        sheet.cell(row=row, column=6, value=entry["runs_per_case"])
        row += 1

    row += 1
    row = header(sheet, row, ["4.3 Agent 1 confidence", "n", "Mean", "Median", "Min", "Max"])
    for dim in dims:
        entry = rel["agent1_confidence_4_3"][dim]
        for column, value in enumerate(
            [dim, entry["n"], entry["mean"], entry["median"], entry["min"], entry["max"]],
            start=1,
        ):
            sheet.cell(row=row, column=column, value=value)
        row += 1

    row += 1
    row = caveat(sheet, row, f"4.2 EVIDENCE RATE - {rel['evidence_caveat'].upper()}")
    row = caveat(
        sheet,
        row,
        "DO NOT QUOTE THE FIGURES BELOW WITHOUT THIS FLAG. Agent 1 produced "
        "verbatim quotes that were not always probative for the dimension they "
        "were attached to, so stated_in_brief is inflated.",
    )
    row = header(sheet, row, ["4.2 Evidence rate", "n", "of", "%", "Status"])
    for dim in dims:
        entry = rel["agent1_evidence_rate_4_2"][dim]
        sheet.cell(row=row, column=1, value=dim)
        write_rate(sheet, row, 2, entry)
        cell = sheet.cell(row=row, column=5, value=entry["caveat"])
        cell.font = CAVEAT_FONT
        row += 1
    finish(sheet)


def sheet_efficiency(book: Workbook, k: dict[str, Any]) -> None:
    """Tokens, cost and wall clock, with the illustrative-only flag on 7.4."""
    sheet = book.create_sheet("Efficiency")
    eff = k["efficiency"]
    title(sheet, "Efficiency - tokens, cost and wall clock per variant")
    row = 3

    row = header(
        sheet,
        row,
        [
            "Variant",
            "Runs",
            "7.1 Input tokens (total)",
            "7.1 Input (mean)",
            "7.1 Output tokens (total)",
            "7.1 Output (mean)",
            "7.2 Cost USD (total)",
            "7.2 Cost USD (mean)",
            "7.3 Wall clock s (total)",
            "7.3 Wall clock s (mean)",
        ],
    )
    for variant in k["meta"]["variants"]:
        entry = eff["per_variant"][variant]
        for column, value in enumerate(
            [
                variant,
                entry["n_runs"],
                entry["input_tokens_7_1"]["total"],
                entry["input_tokens_7_1"]["mean"],
                entry["output_tokens_7_1"]["total"],
                entry["output_tokens_7_1"]["mean"],
                entry["cost_usd_7_2"]["total"],
                entry["cost_usd_7_2"]["mean"],
                entry["wall_clock_s_7_3"]["total"],
                entry["wall_clock_s_7_3"]["mean"],
            ],
            start=1,
        ):
            sheet.cell(row=row, column=column, value=value)
        row += 1

    row += 1
    cpi = eff["cost_per_improvement_7_4"]
    row = caveat(sheet, row, f"7.4 COST PER IMPROVEMENT - {cpi['note']}")
    row = header(sheet, row, ["7.4 Cost per improvement", "Value"])
    for label, value in [
        (
            "Extra cost, full over minus_critic (USD)",
            cpi["extra_cost_full_over_minus_critic_usd"],
        ),
        ("Improvements observed", cpi["improvements_observed"]),
    ]:
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=value)
        row += 1
    finish(sheet)


def sheet_deferred(book: Workbook, k: dict[str, Any]) -> None:
    """What Campaign 1 deliberately did not compute, and why."""
    sheet = book.create_sheet("Deferred - Not Computed")
    title(sheet, "Deliberately not computed in Campaign 1")
    row = 3
    row = caveat(
        sheet,
        row,
        "These are intentional gaps, not oversights. Each requires data that "
        "Campaign 1 did not produce.",
    )
    row += 1
    row = header(sheet, row, ["KPI", "Why it was not computed"])
    for kpi, why in k["deferred"].items():
        sheet.cell(row=row, column=1, value=kpi)
        sheet.cell(row=row, column=2, value=why)
        row += 1
    finish(sheet, freeze="A7")


def build(kpis: Path, out: Path) -> Path:
    """Write the workbook and return its path."""
    k = json.loads(kpis.read_text(encoding="utf-8"))
    book = Workbook()
    book.remove(book.active)

    sheet_overview(book, k)
    sheet_rq1(book, k)
    sheet_rq2(book, k)
    sheet_rq3(book, k)
    sheet_rq4(book, k)
    sheet_reliability(book, k)
    sheet_efficiency(book, k)
    sheet_deferred(book, k)

    out.parent.mkdir(parents=True, exist_ok=True)
    book.save(out)
    return out


def parse_args(argv: Optional[list[str]] = None) -> argparse.Namespace:
    """Parse the command line."""
    parser = argparse.ArgumentParser(description="Export Campaign 1 KPIs to Excel.")
    parser.add_argument("--kpis", type=Path, default=Path("analysis/kpis.json"))
    parser.add_argument(
        "--out", type=Path, default=Path("analysis/LUXCAL_Campaign1_KPIs.xlsx")
    )
    return parser.parse_args(argv)


def main(argv: Optional[list[str]] = None) -> int:
    """Build the workbook, then read it back to verify it opens."""
    args = parse_args(argv)
    path = build(args.kpis, args.out)

    book = load_workbook(path)
    print(f"wrote {path}\n")
    print(f"{'sheet':<32}{'rows':>6}{'cols':>6}  {'frozen':>8}")
    for name in book.sheetnames:
        sheet = book[name]
        print(
            f"{name:<32}{sheet.max_row:>6}{sheet.max_column:>6}  "
            f"{str(sheet.freeze_panes):>8}"
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
